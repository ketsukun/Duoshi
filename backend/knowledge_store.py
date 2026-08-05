import os
import threading
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


CHINESE_HEADERS = [
    "典故编号",
    "典故名称",
    "典故原文",
    "典故释义",
    "语义标签",
    "典故异名",
    "诗例",
]


class KnowledgeStore:
    def __init__(self, workbook_path):
        self.workbook_path = Path(workbook_path)
        self._write_lock = threading.Lock()

    def _create_workbook(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "用户扩充典故"
        worksheet.append(CHINESE_HEADERS)
        worksheet.freeze_panes = "A2"

        header_fill = PatternFill("solid", fgColor="2F6D5B")
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        column_widths = [12, 20, 48, 48, 24, 24, 42]
        for index, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        return workbook

    def _load_workbook(self):
        if not self.workbook_path.exists():
            return self._create_workbook()

        workbook = load_workbook(self.workbook_path)
        worksheet = workbook.active
        actual_headers = [cell.value for cell in worksheet[1]]
        if actual_headers != CHINESE_HEADERS:
            raise ValueError("用户知识库表头不符合预期，无法安全追加数据。")
        return workbook

    @staticmethod
    def _next_allusion_id(worksheet):
        existing_ids = [
            value
            for (value,) in worksheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if isinstance(value, int)
        ]
        return max(existing_ids, default=0) + 1

    @staticmethod
    def _update_table(worksheet):
        table_reference = f"A1:G{worksheet.max_row}"
        if "UserAllusions" in worksheet.tables:
            worksheet.tables["UserAllusions"].ref = table_reference
            return

        table = Table(displayName="UserAllusions", ref=table_reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    def add(self, entry):
        with self._write_lock:
            workbook = self._load_workbook()
            worksheet = workbook.active
            allusion_id = self._next_allusion_id(worksheet)
            worksheet.append([
                allusion_id,
                entry["allusion_name"],
                entry["source_text"],
                entry["allusion_mean"],
                entry.get("semantic_tags", ""),
                entry.get("allusion_variants", ""),
                entry["poem_example"],
            ])
            self._update_table(worksheet)

            self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
            temporary_path = self.workbook_path.with_suffix(".tmp.xlsx")
            try:
                workbook.save(temporary_path)
                os.replace(temporary_path, self.workbook_path)
            finally:
                if temporary_path.exists():
                    temporary_path.unlink()
                workbook.close()
            return allusion_id

    def count(self):
        with self._write_lock:
            if not self.workbook_path.exists():
                return 0
            workbook = self._load_workbook()
            try:
                return max(0, workbook.active.max_row - 1)
            finally:
                workbook.close()
